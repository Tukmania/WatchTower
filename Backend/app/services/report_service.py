import os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Location constants ─────────────────────────────────────────────────────────
TERMINALS = [
    'TERMINAL 1A',
    'TERMINAL 1B',
    'TERMINAL 1C',
    'TERMINAL 1D',
    'MARINA'
]

SHOPS = [
    'LAOMAI JKIA',
    'LAOMAI MOMBASA',
    'CRAYSON PHARMACY JKIA',
    'CRAYSON PHARMACY KIAMBU ROAD',
    'CRAYSON PHARMACY MOMBASA'
]

# ── Dayshift blocks 07:00 → 18:00 ─────────────────────────────────────────────
DAYSHIFT_BLOCKS = []
for _h in range(7, 18):
    for _m in [0, 30]:
        _end_m = _m + 30
        _end_h = _h
        if _end_m >= 60:
            _end_m -= 60
            _end_h += 1
        DAYSHIFT_BLOCKS.append(
            f"{_h:02d}:{_m:02d} – {_end_h:02d}:{_end_m:02d}"
        )

# ── Nightshift blocks 18:00 → 07:00 ───────────────────────────────────────────
NIGHTSHIFT_BLOCKS = []
for _h in range(18, 24):
    for _m in [0, 30]:
        _end_m = _m + 30
        _end_h = _h
        if _end_m >= 60:
            _end_m -= 60
            _end_h += 1
        if _end_h >= 24:
            _end_h = 0
        NIGHTSHIFT_BLOCKS.append(
            f"{_h:02d}:{_m:02d} – {_end_h:02d}:{_end_m:02d}"
        )
for _h in range(0, 7):
    for _m in [0, 30]:
        _end_m = _m + 30
        _end_h = _h
        if _end_m >= 60:
            _end_m -= 60
            _end_h += 1
        NIGHTSHIFT_BLOCKS.append(
            f"{_h:02d}:{_m:02d} – {_end_h:02d}:{_end_m:02d}"
        )


# ── Style helpers ──────────────────────────────────────────────────────────────
def _thin_side():
    return Side(style='thin')


def _thin_border():
    t = _thin_side()
    return Border(left=t, right=t, top=t, bottom=t)


def _blank_border():
    """No border — used for separator columns."""
    return Border()


def write_cell(ws, row, col, value='',
               bold=False,
               bg=None,
               font_color='000000',
               align='center',
               font_size=9,
               wrap=False,
               border=True,
               italic=False):
    """
    Safely writes to a cell only if it is NOT part of a merge range
    (i.e. not a MergedCell slave). If it is a merged slave, skip silently.
    """
    cell = ws.cell(row=row, column=col)

    # openpyxl marks merged slave cells as MergedCell type
    # Writing value to them raises AttributeError — skip them
    from openpyxl.cell.cell import MergedCell
    if isinstance(cell, MergedCell):
        return cell

    cell.value = value
    cell.font  = Font(
        bold=bold,
        color=font_color,
        size=font_size,
        name='Arial',
        italic=italic
    )
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(
        horizontal=align,
        vertical='center',
        wrap_text=wrap
    )
    cell.border = _thin_border() if border else _blank_border()
    return cell


def write_blank_separator(ws, row, col, bg='E2E8F0'):
    """Writes a blank separator column cell — no border, light grey fill."""
    cell = ws.cell(row=row, column=col)
    from openpyxl.cell.cell import MergedCell
    if isinstance(cell, MergedCell):
        return
    cell.value = ''
    cell.fill  = PatternFill('solid', fgColor=bg)
    cell.border = _blank_border()


# ── Main entry point ───────────────────────────────────────────────────────────
def generate_excel_report(events: list, label: str, filename_base: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CCTV Report"

    # ── Aggregate events → data[block][location][metric] = count ──────────────
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for e in events:
        block    = e.get('time_block', '')
        location = e.get('location', '')
        subtype  = e.get('subtype', '')
        etype    = e.get('event_type', '')
        category = e.get('location_category', '')

        if category == 'terminal':
            if subtype == 'bag_wrap':
                data[block][location]['bag_wrap'] += 1
            elif subtype == 'box_wrap':
                data[block][location]['box_wrap'] += 1
            elif subtype == 'foot_traffic':
                data[block][location]['foot_traffic'] += 1

        elif category == 'shop':
            if etype == 'interaction':
                data[block][location]['interaction'] += 1
            elif etype == 'walkin':
                data[block][location]['walkin'] += 1
            elif etype == 'receipt':
                data[block][location]['receipt'] += 1

    # ── Determine which time blocks are actually present in the data ───────────
    # Only write blocks that fall within the requested interval
    present_blocks = set(data.keys())

    def _filter_blocks(block_list):
        # Keep all defined blocks; zeros will show for empty ones.
        # The caller already filtered events by timestamp, so only relevant
        # time blocks will have non-zero data.
        return block_list

    # ── Write dayshift section ─────────────────────────────────────────────────
    current_row = 1
    current_row = _write_shift_section(
        ws, data, label,
        shift_label='DAYSHIFT',
        blocks=DAYSHIFT_BLOCKS,
        start_row=current_row
    )

    # Two blank rows between sections
    current_row += 2

    # ── Write nightshift section ───────────────────────────────────────────────
    _write_shift_section(
        ws, data, label,
        shift_label='NIGHTSHIFT',
        blocks=NIGHTSHIFT_BLOCKS,
        start_row=current_row
    )

    # ── Freeze the top rows of first section so headers stay visible ───────────
    ws.freeze_panes = 'B6'

    # ── Save ───────────────────────────────────────────────────────────────────
    output_dir = os.path.abspath(
        os.path.join(os.path.dirname(__file__), '..', '..', 'reports', 'output')
    )
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f'CCTV_Report_{filename_base}.xlsx')
    wb.save(filepath)
    return filepath


# ── Section writer ─────────────────────────────────────────────────────────────
def _write_shift_section(ws, data, label, shift_label, blocks, start_row):
    """
    Writes one full shift section onto the worksheet.
    Returns the next available row index after the section.

    Column map:
      Col  1        → TIME INTERVAL
      Col  2– 4     → TERMINAL 1A  (bag, box, foot)
      Col  5– 7     → TERMINAL 1B
      Col  8–10     → TERMINAL 1C
      Col 11–13     → TERMINAL 1D
      Col 14–16     → MARINA
      Col 17        → blank separator
      Col 18–20     → LAOMAI JKIA
      Col 21        → blank separator
      Col 22–24     → LAOMAI MOMBASA
      Col 25        → blank separator
      Col 26–28     → CRAYSON PHARMACY JKIA
      Col 29        → blank separator
      Col 30–32     → CRAYSON PHARMACY KIAMBU ROAD
      Col 33        → blank separator
      Col 34–36     → CRAYSON PHARMACY MOMBASA
    """

    # ── Colors ────────────────────────────────────────────────────────────────
    C_TITLE     = '1B2B5A'
    C_COMPANY   = '243352'
    C_TERM_HEAD = '2563EB'
    C_SHOP_HEAD = '0F766E'
    C_SUB_HEAD  = '334155'
    C_TIME_COL  = 'F1F5F9'
    C_EVEN_TERM = 'F8FAFF'
    C_EVEN_SHOP = 'F0FDFA'
    C_WHITE     = 'FFFFFF'
    C_TOTAL     = '1E3A5F'
    C_SEP       = 'E2E8F0'

    LAST_COL           = 36
    TERMINAL_START_COLS = [2, 5, 8, 11, 14]
    SHOP_START_COLS     = [18, 22, 26, 30, 34]
    SEP_COLS            = [17, 21, 25, 29, 33]   # blank separator columns

    r = start_row

    # ── Helper: paint all separator columns in a given row ────────────────────
    def paint_separators(row):
        for sc in SEP_COLS:
            write_blank_separator(ws, row, sc, C_SEP)

    # ── ROW 1 — Shift label ───────────────────────────────────────────────────
    ws.merge_cells(
        start_row=r, start_column=1,
        end_row=r,   end_column=LAST_COL
    )
    write_cell(ws, r, 1, shift_label,
               bold=True, bg=C_TITLE, font_color=C_WHITE,
               font_size=12, align='left')
    ws.row_dimensions[r].height = 20
    r += 1

    # ── ROW 2 — Company + date ────────────────────────────────────────────────
    ws.merge_cells(
        start_row=r, start_column=1,
        end_row=r,   end_column=LAST_COL
    )
    write_cell(
        ws, r, 1,
        f'KIJANI LIMITED                                  {label}',
        bold=True, bg=C_COMPANY, font_color=C_WHITE,
        font_size=10, align='left'
    )
    ws.row_dimensions[r].height = 16
    r += 1

    # ── ROW 3 — Group headers (terminal names + shop names) ───────────────────
    # TIME INTERVAL cell merges rows 3,4,5 into one tall cell
    ws.merge_cells(
        start_row=r, start_column=1,
        end_row=r + 2, end_column=1
    )
    write_cell(ws, r, 1, 'TIME INTERVAL',
               bold=True, bg=C_SUB_HEAD, font_color=C_WHITE,
               font_size=8, wrap=True)

    # Terminal group headers — each spans 3 columns across row 3 only
    terminal_labels = ['TERMINAL 1A', 'TERMINAL 1B', 'TERMINAL 1C',
                       'TERMINAL 1D', 'MARINA']
    for i, label in enumerate(terminal_labels):
        sc = TERMINAL_START_COLS[i]
        ws.merge_cells(
            start_row=r, start_column=sc,
            end_row=r,   end_column=sc + 2
        )
        write_cell(ws, r, sc, label,
                   bold=True, bg=C_TERM_HEAD, font_color=C_WHITE, font_size=8)

    # Shop group headers — each spans 3 columns across row 3 only
    for i, shop in enumerate(SHOPS):
        sc = SHOP_START_COLS[i]
        ws.merge_cells(
            start_row=r, start_column=sc,
            end_row=r,   end_column=sc + 2
        )
        write_cell(ws, r, sc, shop,
                   bold=True, bg=C_SHOP_HEAD, font_color=C_WHITE,
                   font_size=7, wrap=True)

    paint_separators(r)
    ws.row_dimensions[r].height = 24
    r += 1

    # ── ROW 4 — "FT" label row ────────────────────────────────────────────────
    # Column 1 is a merged slave here — write_cell skips it safely
    for i in range(len(TERMINALS)):
        sc = TERMINAL_START_COLS[i]
        write_cell(ws, r, sc,   '', bg=C_TERM_HEAD, font_color=C_WHITE, font_size=7)
        write_cell(ws, r, sc+1, '', bg=C_TERM_HEAD, font_color=C_WHITE, font_size=7)
        write_cell(ws, r, sc+2, '', bg=C_TERM_HEAD, font_color=C_WHITE, font_size=7)

    for i in range(len(SHOPS)):
        sc = SHOP_START_COLS[i]
        write_cell(ws, r, sc,   'FT', bold=True, bg=C_SHOP_HEAD,
                   font_color=C_WHITE, font_size=7)
        write_cell(ws, r, sc+1, '',   bg=C_SHOP_HEAD,
                   font_color=C_WHITE, font_size=7)
        write_cell(ws, r, sc+2, '',   bg=C_SHOP_HEAD,
                   font_color=C_WHITE, font_size=7)

    paint_separators(r)
    ws.row_dimensions[r].height = 12
    r += 1

    # ── ROW 5 — Sub-headers BAG/BOX/FOOT and INT/WLK/REC ─────────────────────
    # Column 1 is still a merged slave — write_cell skips it safely
    sub_terminal = ['BAG WRAP', 'BOX WRAP', 'FOOT TRAFFIC']
    for i in range(len(TERMINALS)):
        sc = TERMINAL_START_COLS[i]
        for j, lbl in enumerate(sub_terminal):
            write_cell(ws, r, sc + j, lbl,
                       bold=True, bg=C_SUB_HEAD, font_color=C_WHITE,
                       font_size=7, wrap=True)

    sub_shop = ['INTERACTION', 'WALK IN', 'RECEIPTS']
    for i in range(len(SHOPS)):
        sc = SHOP_START_COLS[i]
        for j, lbl in enumerate(sub_shop):
            write_cell(ws, r, sc + j, lbl,
                       bold=True, bg=C_SUB_HEAD, font_color=C_WHITE,
                       font_size=7, wrap=True)

    paint_separators(r)
    ws.row_dimensions[r].height = 22
    r += 1

    # ── DATA ROWS ─────────────────────────────────────────────────────────────
    # Every block in the list gets a row — zeros if no events were logged
    totals = defaultdict(lambda: defaultdict(int))

    for idx, block in enumerate(blocks):
        bg_term = C_EVEN_TERM if idx % 2 == 0 else C_WHITE
        bg_shop = C_EVEN_SHOP if idx % 2 == 0 else C_WHITE

        # Time interval
        write_cell(ws, r, 1, block,
                   bg=C_TIME_COL, font_size=8, align='center')

        # Terminal columns — always write 0 when no data
        for i, terminal in enumerate(TERMINALS):
            sc   = TERMINAL_START_COLS[i]
            bag  = data[block][terminal].get('bag_wrap', 0)
            box  = data[block][terminal].get('box_wrap', 0)
            foot = data[block][terminal].get('foot_traffic', 0)

            write_cell(ws, r, sc,   bag,  bg=bg_term, font_size=9, align='center')
            write_cell(ws, r, sc+1, box,  bg=bg_term, font_size=9, align='center')
            write_cell(ws, r, sc+2, foot, bg=bg_term, font_size=9, align='center')

            totals[terminal]['bag_wrap']     += bag
            totals[terminal]['box_wrap']     += box
            totals[terminal]['foot_traffic'] += foot

        # Shop columns — always write 0 when no data
        for i, shop in enumerate(SHOPS):
            sc          = SHOP_START_COLS[i]
            interaction = data[block][shop].get('interaction', 0)
            walkin      = data[block][shop].get('walkin', 0)
            receipt     = data[block][shop].get('receipt', 0)

            write_cell(ws, r, sc,   interaction, bg=bg_shop, font_size=9, align='center')
            write_cell(ws, r, sc+1, walkin,      bg=bg_shop, font_size=9, align='center')
            write_cell(ws, r, sc+2, receipt,     bg=bg_shop, font_size=9, align='center')

            totals[shop]['interaction'] += interaction
            totals[shop]['walkin']      += walkin
            totals[shop]['receipt']     += receipt

        paint_separators(r)
        ws.row_dimensions[r].height = 14
        r += 1

    # ── TOTAL ROW ─────────────────────────────────────────────────────────────
    write_cell(ws, r, 1, 'TOTAL',
               bold=True, bg=C_TOTAL, font_color=C_WHITE, font_size=9)

    for i, terminal in enumerate(TERMINALS):
        sc = TERMINAL_START_COLS[i]
        write_cell(ws, r, sc,   totals[terminal]['bag_wrap'],
                   bold=True, bg=C_TOTAL, font_color=C_WHITE, font_size=9)
        write_cell(ws, r, sc+1, totals[terminal]['box_wrap'],
                   bold=True, bg=C_TOTAL, font_color=C_WHITE, font_size=9)
        write_cell(ws, r, sc+2, totals[terminal]['foot_traffic'],
                   bold=True, bg=C_TOTAL, font_color=C_WHITE, font_size=9)

    for i, shop in enumerate(SHOPS):
        sc = SHOP_START_COLS[i]
        write_cell(ws, r, sc,   totals[shop]['interaction'],
                   bold=True, bg=C_TOTAL, font_color=C_WHITE, font_size=9)
        write_cell(ws, r, sc+1, totals[shop]['walkin'],
                   bold=True, bg=C_TOTAL, font_color=C_WHITE, font_size=9)
        write_cell(ws, r, sc+2, totals[shop]['receipt'],
                   bold=True, bg=C_TOTAL, font_color=C_WHITE, font_size=9)

    paint_separators(r)
    ws.row_dimensions[r].height = 16
    r += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 15
    for col_idx in range(2, 17):
        ws.column_dimensions[get_column_letter(col_idx)].width = 9
    for col_idx in SEP_COLS:
        ws.column_dimensions[get_column_letter(col_idx)].width = 2
    for sc in SHOP_START_COLS:
        for offset in range(3):
            ws.column_dimensions[get_column_letter(sc + offset)].width = 11

    return r